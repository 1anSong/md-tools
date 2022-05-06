from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

savedir = 'image_xcsong/' "存放图片的路径"

wb = load_workbook("work.xlsx") "加载excel文件"
ws = wb["1"] "加载名字为1的工作表"

image_loader = SheetImageLoader(ws) "加载图片"

lumo=ws['B1'].value "读取B1单元格"
humo=ws['C1'].value "读取C1单元格"

for i in range(2,31):
    '''
    name_lumo = ws['A'+str(i)].value.split('/')[-1]+'-'+lumo
    image=image_loader.get('B'+str(i))
    image.save(savedir+name_lumo+".tiff")
    '''
    name_humo = ws['A'+str(i)].value.split('/')[-1]+'-'+humo
    image=image_loader.get('C'+str(i))
    image.save(savedir+name_humo+'.tiff')

